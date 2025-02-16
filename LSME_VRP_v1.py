import os
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd
import math
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from string import ascii_uppercase
from datetime import datetime

vrp_folder_path = r"C:\Users\jerem\Desktop\Studienassistenz\LSME\VRP XML"
excel_file = r"C:\Users\jerem\Desktop\Studienassistenz\LSME\LSME_Auswertung.xlsx"

def calculate_semitones(pitch_low, pitch_high):
    return round(12 * math.log2(pitch_high / pitch_low))

def calculate_dynamic(volume_low, volume_high):
    return (volume_high - volume_low)

def process_xml_files_in_folder (vrp_folder_path, excel_file):

    workbook = openpyxl.Workbook()
    vrp_sheet = workbook.active
    vrp_sheet.title = ("VRP")
    vowels_sheet = workbook.create_sheet(title="Vokale")
    ltas_speech_sheet = workbook.create_sheet(title="Sprechtext")
    ltas_singing_sheet = workbook.create_sheet(title="Summertime")
    counting_sheet = workbook.create_sheet(title="Anzahl der Messungen")

    vrp_sheet.cell(row=1, column=1).value = "Proband_in"
    vrp_sheet.cell(row=2, column=1).value = "Datum"

    xml_files = [f for f in os.listdir(vrp_folder_path) if f.endswith(".xml")]

    if not xml_files:
        print("Keine XML-Dateien im Ordner gefunden.")
        return

    data_for_sorting = []

    for col_idx, xml_file in enumerate(xml_files, start=2):
        xml_path = os.path.join(vrp_folder_path, xml_file)
        print(f"Verarbeite Datei: {xml_path}")

        # XML-Datei einlesen und parsen
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Session-, Analyse- und Datumsinformationen extrahieren
        session = root.find(".//session")
        data = root.find(".//data")
        date_element = root.find(".//date")  # Datum aus XML

        if session is None or data is None:
            print(f"Warnung: Session- oder Analyse-Daten fehlen in Datei: {xml_file}")
            continue

        # Client Number auslesen und "lwintern" entfernen
        client_number = session.find("client_number").text.replace("lwintern", "")
        client_number = int(client_number)  # In echte Zahl umwandeln
        vrp_sheet.cell(row=1, column=col_idx).value = client_number  # Client Number in Kopfzeile

        # Datum aus dem <date>-Tag eintragen
        if date_element is not None:
            vrp_sheet.cell(row=2, column=col_idx).value = date_element.text  # Datum direkt aus dem XML übernehmen
        else:
            vrp_sheet.cell(row=2, column=col_idx).value = "Unbekannt"

        # Letzter Buchstabe des Dateinamens in Zeile 3
        last_char = xml_file[-5]  # Vor `.xml`
        vrp_sheet.cell(row=3, column=col_idx).value = last_char


        # Schleife durch alle Zeilen der ersten Spalte (A)
        for row in range(1, vrp_sheet.max_row + 1):
            if vrp_sheet.cell(row=row, column=1).value == "m/f":
                break  # Schleife beenden, wenn der Wert gefunden wurde

        # Analyse-Daten schreiben
        row_offset = 0  # Variable, um zusätzliche Zeilen dynamisch einzufügen
        for row_idx, child in enumerate(data, start=3):
            # Erste Spalte: Typen (nur einmal schreiben)
            if col_idx == 2:  # Nur für die erste Datei
                vrp_sheet.cell(row=row_idx, column=1).value = child.tag

            # Werte in die entsprechende Spalte
            vrp_sheet.cell(row=row_idx, column=col_idx).value = (
                float(child.text) if child.text.replace(".", "").isdigit() else child.text
            )

    vrp_sheet.column_dimensions['A'].width = 25

    singing_pitch_max_row = None
    for row in range(1, vrp_sheet.max_row + 1):
        if vrp_sheet.cell(row=row, column=1).value == "singing_pitch_max":
            singing_pitch_max_row = row
            break
    norm_profile_coverage_row = None
    for row in range(1, vrp_sheet.max_row + 1):
        if vrp_sheet.cell(row=row, column=1).value == "norm_profile_coverage":
            norm_profile_coverage_row = row
            break

    num_columns = vrp_sheet.max_column

#Gesangs-Tonumfang berechnen und in einer neuen Zeile hinzufügen
    vrp_sheet.insert_rows(norm_profile_coverage_row + 1)
    for col in range(2, num_columns + 1):
        singing_pitch_high_value = vrp_sheet.cell(row=singing_pitch_max_row, column=col).value
        singing_pitch_low_value = vrp_sheet.cell(row=singing_pitch_max_row + 1, column=col).value
        if isinstance(singing_pitch_low_value, (int, float)) and isinstance(singing_pitch_high_value, (int, float)):
            semitones = calculate_semitones(singing_pitch_low_value, singing_pitch_high_value)
        else:
            semitones = None  # Falls ein Wert ungültig ist, None setzen
        vrp_sheet.cell(row=norm_profile_coverage_row + 1, column=col).value = semitones
    vrp_sheet.cell(row=norm_profile_coverage_row +1, column=1).value = "Tonumfang"

#Gesangs-Dynamik berechnen und in einer neuen Zeile hinzufügen
    vrp_sheet.insert_rows(norm_profile_coverage_row + 2)
    for col in range(2, num_columns + 1):
        singing_dynamic_high_value = vrp_sheet.cell(row=singing_pitch_max_row + 3, column=col).value
        singing_dynamic_low_value = vrp_sheet.cell(row=singing_pitch_max_row + 4, column=col).value
        if isinstance(singing_dynamic_low_value, (int, float)) and isinstance(singing_dynamic_high_value, (int, float)):
            singing_dynamic = calculate_dynamic(singing_dynamic_low_value, singing_dynamic_high_value)
        else:
            singing_dynamic = None
        vrp_sheet.cell(row=norm_profile_coverage_row + 2, column=col).value = singing_dynamic
    vrp_sheet.cell(row=norm_profile_coverage_row + 2, column=1).value = "Dynamikumfang"


    # Entferne die Leerzeichen
    for row in range(1, vrp_sheet.max_row + 1):
        for col in range(2, vrp_sheet.max_column + 1):
            cell_value = vrp_sheet.cell(row=row, column=col).value
            if isinstance(cell_value, str):
                vrp_sheet.cell(row=row, column=col).value = cell_value.replace(" ", "")


    # Alle Zellen ab Spalte 2 und Zeile 5 in Zahlen umwandeln, falls möglich
    for row in range(5, vrp_sheet.max_row + 1):
        for col in range(2, vrp_sheet.max_column + 1):
            cell = vrp_sheet.cell(row=row, column=col)
            try:
                if cell.value is not None:
                    cell.value = float(cell.value)
            except (ValueError, TypeError):
                pass

    def sort_excel_columns(vrp_sheet, primary_sort_row=1, secondary_sort_row=2):
        columns_data = []
        for col in range(2, vrp_sheet.max_column + 1):
            primary_value = vrp_sheet.cell(row=primary_sort_row, column=col).value
            secondary_value = vrp_sheet.cell(row=secondary_sort_row, column=col).value
            columns_data.append((primary_value, secondary_value, col))

        # Spalten basierend auf den Werten sortieren
        # Zuerst nach `primary_value`, dann nach `secondary_value`
        columns_data.sort(
            key=lambda x: (
                x[0] is None,  # None-Werte ans Ende
                x[0],  # Primäres Sortierkriterium
                x[1] is None,  # None-Werte in der zweiten Zeile ans Ende
                x[1],  # Sekundäres Sortierkriterium
                x[2]  # Originalspaltennummer als Tiebreaker
            )
        )

        # Eine Kopie der aktuellen Tabelle erstellen
        sorted_data = []
        for row in range(1, vrp_sheet.max_row + 1):
            sorted_data.append([vrp_sheet.cell(row=row, column=col).value for _, _, col in columns_data])

        # Die Tabelle mit den sortierten Daten neu schreiben
        for row_idx, row_data in enumerate(sorted_data, start=1):
            for col_idx, value in enumerate(row_data, start=2):  # Neu anordnen ab Spalte 2
                vrp_sheet.cell(row=row_idx, column=col_idx).value = value

        # Überschüssige Spalten löschen (falls nötig)
        while vrp_sheet.max_column > len(columns_data) + 1:
            vrp_sheet.delete_cols(vrp_sheet.max_column)


    # Beispielaufruf
    sort_excel_columns(vrp_sheet, primary_sort_row=1, secondary_sort_row=2)


    #Hinzufügen von A-B-C
    row1 = [cell.value for cell in vrp_sheet[1]]
    row2 = [cell.value for cell in vrp_sheet[2]]

    def extract_year(value):
        try:
            if isinstance(value, datetime):
                return value.year
            elif isinstance(value, (int, float)):
                # Möglicher Excel-Datumswert
                excel_date_origin = datetime(1899, 12, 30)
                return (excel_date_origin + timedelta(days=value)).year
            elif isinstance(value, str):
                return datetime.strptime(value, "%Y/%m/%d").year
            return None
        except (ValueError, TypeError):
            return None
    row2_years = [extract_year(value) for value in row2]

    row3 = []
    for year in row2_years:
        if year is None:
            row3.append("Unbekannt")
        elif year <= 2011:
            row3.append("A")
        elif 2012 <= year <= 2022:
            row3.append("B")
        elif year >= 2023:
            row3.append("C")
            
    for col_num, value in enumerate(row3, start=1):
        vrp_sheet.cell(row=3, column=col_num, value=value)
    vrp_sheet.cell(row=3, column=1).value = "Messreihe"

    # Counting-Blatt erstellen
    num_columns = vrp_sheet.max_column
    num_rows = vrp_sheet.max_row
    person_numbers = set()
    for col in range(2, num_columns + 1):
        value = vrp_sheet.cell(row=1, column=col).value
        if isinstance(value, int):
            person_numbers.add(value)
    counting_sheet.cell(row=1, column=1).value = "Anzahl der Proband_innen"
    counting_sheet.cell(row=1, column=2).value = len(person_numbers)
    counting_sheet.cell(row=2, column=1).value = "Gesamtzahl der Messungen"
    counting_sheet.cell(row=2, column=2).value = num_columns - 1
    counting_sheet.cell(row=3, column=1).value = "Anzahl der A-Messungen"
    counting_sheet.cell(row=4, column=1).value = "Anzahl der B-Messungen"
    counting_sheet.cell(row=5, column=1).value = "Anzahl der C-Messungen"
    a_count = b_count = c_count = 0
    for col in range(2, num_columns + 1):
        value = vrp_sheet.cell(row=3, column=col).value
        if value == "A":
            a_count += 1
        elif value == "B":
            b_count += 1
        elif value == "C":
            c_count += 1
    counting_sheet.cell(row=3, column=2).value = a_count
    counting_sheet.cell(row=4, column=2).value = b_count
    counting_sheet.cell(row=5, column=2).value = c_count


    workbook.save(excel_file)
    print(f"Die Daten aus dem Ordner '{vrp_folder_path}' wurden erfolgreich in '{excel_file}' gespeichert.")

process_xml_files_in_folder(vrp_folder_path, excel_file)






