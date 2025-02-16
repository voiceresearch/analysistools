import os
import subprocess
import soundfile as sf
import pandas as pd
import re
import numpy as np
import matplotlib.pyplot as plt
from scipy.io import wavfile
import warnings
from scipy.io.wavfile import WavFileWarning
import datetime
from datetime import timedelta
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=WavFileWarning)

input_folder = r"C:\Users\jerem\Desktop\Studienassistenz\LSME\vowels_segments_output\test"
output_folder = os.path.join(input_folder, "converted_hd_results")
os.makedirs(output_folder, exist_ok=True)

exe_path = r"C:\Users\jerem\Downloads\hd\hd\hd.exe"
output_excel_path = os.path.join(output_folder, "hdout_with_stats.xlsx")

offset_start = 0.5
offset_end = 0.3
segment_duration = 1.3
audio_extensions = {".wav"}
vowel_endings = ["A", "E", "I", "O", "U"]

# Ergebnisse nach Dateien speichern
data_by_file = {}

def save_segment(data_segment, sample_rate, file_name):
    wavfile.write(file_name, sample_rate, data_segment)

for file_name in os.listdir(input_folder):
    file_path = os.path.join(input_folder, file_name)

    # Überprüft, ob es sich um eine Audiodatei handelt
    if os.path.isfile(file_path) and os.path.splitext(file_name)[1].lower() in audio_extensions:
        try:
            # Segmentierung der Datei
            sample_rate, data = wavfile.read(file_path)
            data = data if len(data.shape) == 1 else data[:, 0]  # Falls stereo, auf mono reduzieren
            db_threshold = 200
            data_abs = np.abs(data)
            active_regions = data_abs > db_threshold

            min_silence_duration = 1
            min_silence_samples = int(min_silence_duration * sample_rate)

            segments = []
            current_start = None

            for i, active in enumerate(active_regions):
                if active and current_start is None:
                    current_start = i
                elif not active and current_start is not None:
                    current_end = i
                    # Überprüfen, ob das Segment lang genug ist
                    if (current_end - current_start) >= int(segment_duration * sample_rate):
                        # Überprüfen, ob Stille zwischen Segmenten ausreichend ist
                        if not segments or current_start - segments[-1][1] >= min_silence_samples:
                            segments.append((current_start, current_end))
                    current_start = None

            valid_segments = []
            used_vowels = 0

            segment_files = []

            for i, (start_idx, end_idx) in enumerate(segments):
                if used_vowels >= len(vowel_endings):
                    break

                segment_start = max(0, start_idx + int(offset_start * sample_rate))
                segment_end = segment_start + int(segment_duration * sample_rate)

                if segment_end <= len(data):
                    segment_data = data[segment_start:segment_end]
                    valid_segments.append((segment_start, segment_end))

                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    segment_file_name = f"{base_name}_{vowel_endings[used_vowels]}.wav"
                    output_file = os.path.join(output_folder, segment_file_name)
                    save_segment(segment_data, sample_rate, output_file)
                    segment_files.append(output_file)
                    used_vowels += 1

            # Diagramm plotten
            plt.figure(figsize=(15, 5))

            time_axis = np.linspace(0, len(data) / sample_rate, num=len(data))
            plt.plot(time_axis, data, label="Original Wellenform", color='grey')

            for segment_start, segment_end in valid_segments:
                plt.axvspan(segment_start / sample_rate, segment_end / sample_rate, color='lightblue', alpha=0.5)

            plt.title(file_name)
            plt.xlabel("Zeit (s)")
            plt.ylabel("Amplitude")
            plt.xlim(0, time_axis[-1])  # Grenzen des Zeitbereichs festlegen
            plt.legend()
            plt.grid(True)
            plt.show()

            for segment_file_path in segment_files:
                try:
                    # Konvertiert Datei zu PCM 16 Bit
                    conv_file_path = os.path.join(output_folder, f"conv_{os.path.basename(segment_file_path)}")
                    data, samplerate = sf.read(segment_file_path)
                    sf.write(conv_file_path, data, samplerate, subtype='PCM_16')
                    print(f"Datei erfolgreich konvertiert: {conv_file_path}")

                    # Analyse mit hd.exe
                    output_file_path = os.path.join(output_folder, f"{os.path.splitext(os.path.basename(segment_file_path))[0]}_hdout.txt")
                    result = subprocess.run(
                        [exe_path, conv_file_path, output_file_path, "info"],
                        capture_output=True,
                        text=True
                    )

                    if result.returncode == 0:
                        print(f"Analyse abgeschlossen für: {segment_file_path}")
                    else:
                        print(f"Fehler bei der Analyse von {segment_file_path}: {result.stderr}")

                    # Liest die Analyseergebnisse und speichert sie
                    if os.path.exists(output_file_path):
                        with open(output_file_path, "r") as file:
                            lines = file.readlines()

                        data_lines = [line.strip() for line in lines if line.strip() and not line.startswith("#")]
                        data_by_file[segment_file_path] = []

                        for line in data_lines:
                            match = re.match(r"(\d+)\s+[\d\.s\-]+\s+(.*)", line)
                            if match:
                                key_values = match.group(2)
                                for key, value in re.findall(r"(\w+)\s([\d\.\-]+)", key_values):
                                    data_by_file[segment_file_path].append({"Kategorie": key, "Wert": float(value)})
                except Exception as e:
                    print(f"Ein Fehler ist aufgetreten bei der Verarbeitung von Segment {segment_file_path}: {e}")


        except Exception as e:
            print(f"Ein Fehler ist aufgetreten bei Datei {file_name}: {e}")

# Datenverarbeitung und Statistikberechnung
try:
    rows = []

    categories = sorted(set(entry["Kategorie"] for file_data in data_by_file.values() for entry in file_data))

    file_segments = {}
    for segment_file in data_by_file.keys():
        base_name = os.path.basename(segment_file).split("_")[0]
        if base_name not in file_segments:
            file_segments[base_name] = []
        file_segments[base_name].append(segment_file)

    # Definiert die Spaltenüberschriften
    column_names = ["Kategorie"] + list(file_segments.keys())

    data_dict = {category: {base_name: "" for base_name in file_segments.keys()} for category in categories}

    for base_name, segments in file_segments.items():
        for category in categories:
            all_values = [entry["Wert"] for segment in segments for entry in data_by_file[segment] if entry["Kategorie"] == category]
            if all_values:
                mean_value = np.mean(all_values)
                std_value = np.std(all_values)
                data_dict[category][base_name] = f'Mean: {mean_value:.2f}, Std: {std_value:.2f}'

    for category, file_values in data_dict.items():
        mean_row = [f"{category} - Mean"] + [file_values[file].split(", ")[0].replace("Mean: ", "") for file in
                                             column_names[1:]]
        std_row = [f"{category} - Std"] + [file_values[file].split(", ")[1].replace("Std: ", "") for file in
                                           column_names[1:]]

        rows.append(mean_row)
        rows.append(std_row)

    date_row = ["Datum"]
    for base_name in column_names[1:]:
        orig_path = os.path.join(input_folder, base_name + ".wav")
        if not os.path.exists(orig_path):
            orig_path = os.path.join(input_folder, base_name + ".WAV")
        if os.path.exists(orig_path):
            modification_date = datetime.datetime.fromtimestamp(os.path.getmtime(orig_path)).strftime("%Y/%m/%d")
            date_row.append(modification_date)
        else:
            date_row.append("N/A")

    # DataFrame erstellen und speichern
    final_df = pd.DataFrame(rows, columns=column_names)
    final_df.to_excel(output_excel_path, index=False)

    print(f"Ergebnisse in {output_excel_path} gespeichert.")

    def extract_year(value):
        try:
            if isinstance(value, str):
                return datetime.datetime.strptime(value, "%Y/%m/%d").year
            return None
        except (ValueError, TypeError):
            return None

    group_row = ["Messreihe"]
    for date_value in date_row[1:]:
        year = extract_year(date_value)
        if year is None:
            group_row.append("Unbekannt")
        elif year <= 2011:
            group_row.append("A")
        elif 2012 <= year <= 2022:
            group_row.append("B")
        elif year >= 2023:
            group_row.append("C")

    rows.insert(0, group_row)
    rows.insert(0, date_row)

    final_df = pd.DataFrame(rows, columns=column_names)
    final_df.to_excel(output_excel_path, index=False)

    # Temporäre Dateien löschen (außer Excel)
    temp_files = [f for f in os.listdir(output_folder) if
            (f.endswith(".wav") or f.endswith(".txt")) and f != os.path.basename(output_excel_path)]

    for temp_file in temp_files:
        temp_file_path = os.path.join(output_folder, temp_file)
        if os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except Exception as e:
                print(f"Fehler beim Löschen von {temp_file}: {e}")
        else:
            print(f"Datei nicht gefunden: {temp_file_path}")

    wb = load_workbook(output_excel_path)
    ws = wb.active
    for row in range(3, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value is not None:
                    cell.value = float(cell.value)
            except (ValueError, TypeError):
                pass
    ws.column_dimensions['A'].width = 20
    wb.save(output_excel_path)


except Exception as e:
    print(f"Ein Fehler ist aufgetreten: {e}")

print(f"Verarbeitung abgeschlossen. Ergebnisse im Ordner: {output_folder}")